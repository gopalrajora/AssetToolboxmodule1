import sys
import qdarkstyle
import os
from datetime import datetime as dt
import shutil
from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2.QtWidgets import *
from PySide2 import *
import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.utils import *
from PyQt5.QtCore import pyqtSignal, pyqtSlot
import json
import pandas as pd
import numpy as np
from utils import *
from main_cluster_maker import *
from pathlib import Path
region_coordinates = {'asia': [34.0479, 100.6197], 'europe': [53.0000, 9.0000], 'africa': [9.1021, 18.2812], 'usa': [44.5000, -89.500]}

class ChecklistDialog(QDialog):

    def __init__(
            self,
            name,
            stringlist=None,
            checked=False,
            icon=None,
            parent=None,
    ):
        super(ChecklistDialog, self).__init__(parent)

        self.name = name
        self.icon = icon
        self.model = QStandardItemModel()
        self.listView = QListView()

        for string in stringlist:
            item = QStandardItem(string)
            item.setCheckable(True)
            check = (Qt.Checked if checked else Qt.Unchecked)
            item.setCheckState(check)
            self.model.appendRow(item)

        self.listView.setModel(self.model)
        self.listView.setCurrentIndex(self.model.index(0, 0))

        self.okButton = QPushButton('OK')
        self.cancelButton = QPushButton('Cancel')

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(self.okButton)
        hbox.addWidget(self.cancelButton)

        vbox = QVBoxLayout(self)
        vbox.addWidget(self.listView)
        vbox.addLayout(hbox)

        self.setWindowTitle(self.name)
        if self.icon:
            self.setWindowIcon(self.icon)

        self.okButton.clicked.connect(self.onAccepted)
        self.cancelButton.clicked.connect(self.reject)

    def onAccepted(self):
        if self.listView.selectedIndexes().__len__() != 1:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText('Please select only one field')
            msg.setWindowTitle("Error")
            msg.exec_()
            return
        self.choices = [self.listView.selectedIndexes()[0].data()]
        if self.choices.__len__() != 1:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText('Please select only one field')
            msg.setWindowTitle("Error")
            msg.exec_()
        else:
            self.accept()


class LatLongDialog(QDialog):

    def __init__(
            self,
            name,
            columns
    ):
        super(LatLongDialog, self).__init__()

        self.name = name
        self.columns = columns
        self.okButton = QPushButton('OK')
        self.cancelButton = QPushButton('Cancel')

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(self.okButton)
        hbox.addWidget(self.cancelButton)
        
        verticalLayoutRegin = QVBoxLayout(self)
        verticalLayoutRegin.setAlignment(Qt.AlignTop)
        verticalLayoutRegin.setObjectName(u"verticalLayoutRegion")
        
        
        indexLabel = QLabel(self)
        indexLabel.setText("Select The Identifier Column")
        verticalLayoutRegin.addWidget(indexLabel)

        self.indexCB = QComboBox()
        self.indexCB.addItems(self.columns)
        
        verticalLayoutRegin.addWidget(self.indexCB)


        latLabel = QLabel(self)
        latLabel.setText("Select The Lattitude Column")
        verticalLayoutRegin.addWidget(latLabel)

        self.latCB = QComboBox()
        self.latCB.addItems(self.columns)
        
        verticalLayoutRegin.addWidget(self.latCB)


        longLabel = QLabel(self)
        longLabel.setText("Select The Longitude Column")
        verticalLayoutRegin.addWidget(longLabel)

        self.longCB = QComboBox()
        self.longCB.addItems(self.columns)
        
        verticalLayoutRegin.addWidget(self.longCB)

        verticalLayoutRegin.addLayout(hbox)
        

        self.setWindowTitle(self.name)
        

        self.okButton.clicked.connect(self.onAccepted)
        self.cancelButton.clicked.connect(self.reject)

    def onAccepted(self):
        self.identifier = self.indexCB.currentText()
        self.latColumn = self.latCB.currentText()
        self.longColumn = self.longCB.currentText()
        
        self.accept()


class Example(QWidget):

    def __init__(self):
        super().__init__()

        self.initUI()

        self.selectedVariables = ""

    def initUI(self):
        self.totalVLayout = QVBoxLayout(self)
        self.totalVLayout.setObjectName(u"totalVLayout")
        self.totalVLayout.setContentsMargins(20, 20, 20, 20)
        self.totalVLayout.setSpacing(15)

        self.logoTitleHLayout = QHBoxLayout()
        self.logoTitleHLayout.setSpacing(10)
        self.logoTitleHLayout.setObjectName(u"fileBrowse1HLayout")

        self.lblLogo = QLabel(self)
        pixmap = QPixmap('../HTML_ASSETS/resources/logo_white.png')
        self.lblLogo.setPixmap(pixmap)
        self.lblLogo.setScaledContents(True)
        self.lblLogo.setMaximumHeight(50)
        self.lblLogo.setMaximumWidth(300)
        self.logoTitleHLayout.addWidget(self.lblLogo)

        self.lblTitle = QLabel(self)
        self.lblTitle.setText("  Tool for the definition of condition indicators")
        self.lblTitle.setAlignment(Qt.AlignCenter)
        self.lblTitle.setFont(QFont("Arial", 16))
        self.logoTitleHLayout.addWidget(self.lblTitle)

        self.totalVLayout.addLayout(self.logoTitleHLayout)

        self.VEntryLayout = QVBoxLayout(self)
        self.VEntryLayout.setObjectName(u"VEntryLayout")
       
        self.dataEntryList = []

        self.addMoreEntries()
        self.totalVLayout.addLayout(self.VEntryLayout)

        self.addMoreEntryBtn = QPushButton(self)
        self.addMoreEntryBtn.setObjectName(u"addMoreEntryBtn")
        self.addMoreEntryBtn.setText("+ Add")
        self.addMoreEntryBtn.clicked.connect(self.addMoreEntries)
        self.totalVLayout.addWidget(self.addMoreEntryBtn)


        regionCoordinatesLayout = QGridLayout()
        
        verticalLayoutRegin = QVBoxLayout()
        verticalLayoutRegin.setAlignment(Qt.AlignTop)
        verticalLayoutRegin.setObjectName(u"verticalLayoutRegion")
        regionLabel = QLabel(self)
        regionLabel.setText("Select The Region")
        verticalLayoutRegin.addWidget(regionLabel)

        self.regionCB = QComboBox()
        self.regionCB.addItems(["USA", 'Asia', 'Europe', 'Africa'])
        
        self.regionCB.currentIndexChanged.connect(self.regionSelectionChanged)

        verticalLayoutRegin.addWidget(self.regionCB)

        regionCoordinatesLayout.addLayout(verticalLayoutRegin,0,0)




        coordinateFileLayout = QVBoxLayout()
        coordinateFileLayout.setSpacing(10)
        coordinateFileLabel = QLabel(self)
        coordinateFileLabel.setText(" Select the Coordinates File")
        coordinateFileLayout.addWidget(coordinateFileLabel)

        coordinateFileBrowseHLayout = QHBoxLayout()
        coordinateFileBrowseHLayout.setSpacing(10)
        coordinateFileBrowseHLayout.setObjectName(u"coordinateFileBrowseHLayout")

        self.coordinateFilePath = QLineEdit(self)
        self.coordinateFilePath.setObjectName(u"coordinateFilePath")
        self.coordinateFilePath.setReadOnly(True)

        coordinateFileBrowseHLayout.addWidget(self.coordinateFilePath)

        coordinateFileBrowseBtn = QPushButton(self)
        coordinateFileBrowseBtn.setObjectName(u"coordinateFileBrowseBtn")
        coordinateFileBrowseBtn.setMinimumSize(QSize(80, 0))
        coordinateFileBrowseBtn.setText("Browse")
        coordinateFileBrowseBtn.clicked.connect(self.coordinateFileBrowseBtnClicked)

        coordinateFileBrowseHLayout.addWidget(coordinateFileBrowseBtn)
        coordinateFileLayout.addLayout(coordinateFileBrowseHLayout)
        regionCoordinatesLayout.addLayout(coordinateFileLayout,0,1)


        regionCoordinatesLayout.setColumnStretch(0, 1)
        regionCoordinatesLayout.setColumnStretch(1, 2)
        self.totalVLayout.addLayout(regionCoordinatesLayout)






        self.pteLog = QPlainTextEdit(self)
        self.pteLog.setObjectName(u"pteLog")
        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pteLog.sizePolicy().hasHeightForWidth())
        self.pteLog.setSizePolicy(sizePolicy)
        self.pteLog.setMinimumSize(QSize(0, 200))
        self.pteLog.setMaximumSize(QSize(16777215, 400))

        self.totalVLayout.addWidget(self.pteLog)

        self.btnClearSettings = QPushButton(self)
        self.btnClearSettings.setObjectName(u"btnClearSettings")
        self.btnClearSettings.setText("Clear")
        # self.btnClearSettings.setSizePolicy(QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed))
        self.btnClearSettings.clicked.connect(self.clearSettingsFunc)

        self.btnSaveSetting = QPushButton(self)
        self.btnSaveSetting.setObjectName(u"btnSaveSetting")
        self.btnSaveSetting.setText("Save Setting")
        self.btnSaveSetting.clicked.connect(self.btnSaveSetting_clicked)
            

        self.horizontalLayout_31 = QHBoxLayout()
        self.horizontalLayout_31.setSpacing(20)
        self.horizontalLayout_31.setObjectName(u"horizontalLayout_31")
        self.horizontalLayout_31.setContentsMargins(20, -1, 20, -1)

        self.horizontalLayout_31.addWidget(self.btnClearSettings)
        self.horizontalLayout_31.addWidget(self.btnSaveSetting)

        self.horizontalLayout_3 = QHBoxLayout()
        self.horizontalLayout_3.setSpacing(20)
        self.horizontalLayout_3.setObjectName(u"horizontalLayout_3")
        self.horizontalLayout_3.setContentsMargins(20, -1, 20, -1)

        self.btnGenerate = QPushButton(self)
        self.btnGenerate.setObjectName(u"btnGenerate")
        self.btnGenerate.clicked.connect(self.btnGenerate_clicked)

        self.horizontalLayout_3.addWidget(self.btnGenerate)

        self.btnShow = QPushButton(self)
        self.btnShow.setObjectName(u"btnShow")
        self.btnShow.clicked.connect(self.btnShow_clicked)
        self.horizontalLayout_3.addWidget(self.btnShow)

        self.totalVLayout.addLayout(self.horizontalLayout_31)
        self.totalVLayout.addLayout(self.horizontalLayout_3)

        self.retranslateUi()

        self.configs = []
        self.componentField = ""
        self.variables = []

        self.setGeometry(0, 0, 800, 640)
        self.setWindowTitle('WP 5.1 Tools')
        self.show()

        #self.loadSetting()

    def retranslateUi(self):
        self.btnGenerate.setText(QCoreApplication.translate("Form", u"Generate Result", None))
        self.btnShow.setText(QCoreApplication.translate("Form", u"Show Result", None))

    def addMoreEntries(self):
        entryDict = {"index":""}

        firstVLayout = QVBoxLayout()
        firstVLayout.setSpacing(10)
        firstLabel = QLabel(self)
        firstLabel.setText(" Select the CSV/Excel File")
        firstVLayout.addWidget(firstLabel)

        fileBrowseHLayout = QHBoxLayout()
        fileBrowseHLayout.setSpacing(10)
        fileBrowseHLayout.setObjectName(u"fileBrowseHLayout")

        entryDict['txtFilePath'] = QLineEdit(self)
        entryDict['txtFilePath'].setObjectName(u"txtFilePath")
        entryDict['txtFilePath'].setReadOnly(True)

        fileBrowseHLayout.addWidget(entryDict['txtFilePath'])

        entryDict['btnBrowse'] = QPushButton(self)
        entryDict['btnBrowse'].setObjectName(u"btnBrowse")
        entryDict['btnBrowse'].setMinimumSize(QSize(80, 0))
        entryDict['btnBrowse'].setText("Browse")
        entryDict['formNumber'] = len(self.dataEntryList)
        entryDict['btnBrowse'].clicked.connect(lambda: self.btnBrowse_clicked(entryDict['formNumber']))

        fileBrowseHLayout.addWidget(entryDict['btnBrowse'])

        firstVLayout.addLayout(fileBrowseHLayout)
        self.VEntryLayout.addLayout(firstVLayout)

        selectComponentVariableHLayout = QHBoxLayout()
        selectComponentVariableHLayout.setSpacing(15)
        selectComponentVariableHLayout.setObjectName(u"selectComponentVariableHLayout")

        secondVLayout = QVBoxLayout()
        secondVLayout.setSpacing(10)
        secondLabel = QLabel(self)
        secondLabel.setText("Select the Variables")
        secondVLayout.addWidget(secondLabel)

        entryDict['twDynamic'] = QTreeWidget(self)
        __qtreewidgetitem = QTreeWidgetItem(entryDict['twDynamic'])
        __qtreewidgetitem.setCheckState(0, Qt.Checked)
        __qtreewidgetitem1 = QTreeWidgetItem(__qtreewidgetitem)
        __qtreewidgetitem1.setCheckState(0, Qt.Checked)
        __qtreewidgetitem4 = QTreeWidgetItem(__qtreewidgetitem)
        __qtreewidgetitem4.setCheckState(0, Qt.Checked)
        entryDict['twDynamic'].setObjectName(u"twDynamic")
        entryDict['twDynamic'].header().setVisible(False)
        entryDict['twDynamic'].itemClicked.connect(self.twDynamic_onItemClicked)

        secondVLayout.addWidget(entryDict['twDynamic'])

        selectComponentVariableHLayout.addLayout(secondVLayout)

        verticalLayout = QVBoxLayout()
        verticalLayout.setAlignment(Qt.AlignTop)
        verticalLayout.setObjectName(u"verticalLayout")
        thirdLabel = QLabel(self)
        thirdLabel.setText("Enter the Weights")
        verticalLayout.addWidget(thirdLabel)

        entryDict['dimensionLineEdit'] = QLineEdit(self)
        entryDict['dimensionLineEdit'].setObjectName(u"dimensionEdit")

        verticalLayout.addWidget(entryDict['dimensionLineEdit'])


        fourthLabel = QLabel(self)
        fourthLabel.setText("Select Dimension")
        verticalLayout.addWidget(fourthLabel)


        entryDict['newColumnEdit'] = QComboBox(self)
        entryDict['newColumnEdit'].setObjectName(u"columnEdit")
        entryDict['newColumnEdit'].addItems(["Economic Impact", 'Health Indexes', 'Life Assessment', 'Maintenance Stratgy'])

        verticalLayout.addWidget(entryDict['newColumnEdit'])

        selectComponentVariableHLayout.addLayout(verticalLayout)

        self.VEntryLayout.addLayout(selectComponentVariableHLayout)
        self.VEntryLayout.addStretch()
        self.dataEntryList.append(entryDict)


    def regionSelectionChanged(self,i):
	
      print (f"Current index: {i},selection changed: {self.regionCB.currentText()}")

    
    @pyqtSlot(QTreeWidgetItem, int)
    def twDynamic_onItemClicked(self, it, col):
        item_count = it.childCount()
        for i in range(item_count):
            child_item = it.child(i)
            if child_item.isDisabled():
                continue
            child_item.setCheckState(0, it.checkState(0))
            for j in range(child_item.childCount()):
                item = child_item.child(j)
                if item.isDisabled():
                    continue
                item.setCheckState(0, it.checkState(0))

    @pyqtSlot()
    def clearSettingsFunc(self):
        # if os.path.isfile("./config.cfg"):
        #     os.remove("./config.cfg")
        self.configs = []
        for i in range(self.dataEntryList.__len__()):
            self.dataEntryList[i]['txtFilePath'].clear()
            self.clearDynamicTree(i)
            self.dataEntryList[i]['dimensionLineEdit'].clear()
            

        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Clear Settings")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nClear Setting")

    def closeEvent(self, event):
        pass
        # if os.path.isfile("./config.cfg"):
        #     os.remove("./config.cfg")


    def coordinateFileBrowseBtnClicked(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "Open File", "", "CSV Files (*.csv)")
        if filePath:
            self.coordinateFilePath.setText(filePath)

            filename, file_extension = os.path.splitext(filePath)
            print(file_extension)
            if file_extension == ".csv":
                self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file...")

                csv_reader = pd.read_csv(filePath)
                headers = csv_reader.columns

                headerTexts = []

                for header in headers:
                    if header != "":
                        headerTexts.append(header)
                

                latLongDialog = LatLongDialog('Select Latitute and Longitude key names', headerTexts)
                if latLongDialog.exec_() == QDialog.Accepted:
                    self.coordinatesIdentifier = latLongDialog.identifier
                    self.latColumn = latLongDialog.latColumn
                    self.longColumn = latLongDialog.longColumn
                self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file completed")
        else:
            return
        
    
    @pyqtSlot()
    def btnBrowse_clicked(self, btnNumber):
        entryDict = self.dataEntryList[btnNumber]
        qfd = QFileDialog()
        path = ""
        filter = "Excel Files(*.xls *.xlsx *.xlsm *.csv)"
        title = "Open file"
        f = QFileDialog.getOpenFileName(qfd, title, path, filter)
        print(f)
        if f[0] != "":
            entryDict['txtFilePath'].setText(f[0])
        else:
            return
        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Browse file")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nBrowse file")

        filePath = entryDict['txtFilePath'].text()
        self.clearDynamicTree(btnNumber)

        filename, file_extension = os.path.splitext(filePath)
        print(file_extension)
        if file_extension == ".csv":
            self.readCSVFile(filePath, btnNumber)
        elif file_extension == ".xls":
            self.readXLSile(filePath, btnNumber)
        else:
            self.readXLSXile(filePath, btnNumber)

    @pyqtSlot()
    def btnSaveSetting_clicked(self):
        self.configs = []
        for n in range(self.dataEntryList.__len__()):
            dataFrame = self.dataEntryList[n]
            selectAllItem = dataFrame['twDynamic'].topLevelItem(0)
            variablesItem = dataFrame['twDynamic'].topLevelItem(0)

            variables = []
           

            
            for i in range(variablesItem.childCount()):
                if variablesItem.child(i).checkState(0) == Qt.Checked:
                    print(variablesItem.child(i).text(0))
                    variables.append(variablesItem.child(i).text(0))

            
            self.configs.append({
                "config": dataFrame['newColumnEdit'].currentText(),
                "path": dataFrame['txtFilePath'].text(),
                "index": dataFrame['index'],
                "variables": variables,
                "weights": dataFrame['dimensionLineEdit'].text().split(','),
            })

        qfd = QFileDialog()
        path = ""
        filter = "Config File(*.cfg)"
        title = "Save Config"
        path = QFileDialog.getSaveFileName(qfd, title, path, filter)
        path = path[0]
        print(path)
        # path = "config.cfg"
        if path != "":
            with open(path, 'w') as f:
                json.dump(self.configs, f)
        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Save Settings")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nSave Setting")

    def loadSetting(self):
        datalen = self.dataEntryList.__len__()
        conflen= self.configs.__len__()
        for _ in range(conflen-datalen if datalen < conflen else 0):
            self.addMoreEntries()

        for i in range(conflen):
            self.dataEntryList[i]['txtFilePath'].setText(self.configs[i]["path"])
            filePath = self.dataEntryList[i]['txtFilePath'].text()

            self.clearDynamicTree(i)

            filename, file_extension = os.path.splitext(filePath)
            print(file_extension)
            self.componentField = self.configs[i]["component1_field"]

            if filePath != "":
                if file_extension == ".csv":
                    self.readCSVFile(filePath)
                elif file_extension == ".xls":
                    self.readXLSile(filePath)
                elif file_extension == ".xlsx":
                    self.readXLSXile(filePath)

            variablesItem = self.dataEntryList[i]['twDynamic'].topLevelItem(0)

            for j in range(variablesItem.childCount()):
                if variablesItem.child(j).text(0) in self.configs[i]["variables"]:
                    variablesItem.child(j).setCheckState(0, Qt.Checked)
                else:
                    variablesItem.child(j).setCheckState(0, Qt.Unchecked)

    def clearDynamicTree(self, formNumber):
        formPart = self.dataEntryList[formNumber]
        # selectAllItem = formPart['twDynamic'].topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = formPart['twDynamic'].topLevelItem(0)

        # for i in reversed(range(componentsItem.childCount())):
        #     componentsItem.removeChild(componentsItem.child(i))
        for i in reversed(range(variablesItem.childCount())):
            variablesItem.removeChild(variablesItem.child(i))
        self.componentField = ""

    def readXLSXile(self, filePath, formNumber):
        formData = self.dataEntryList[formNumber]
        components = []
        componentIndex = 0
        selectAllItem = formData['twDynamic'].topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = formData['twDynamic'].topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file...")
        wb = load_workbook(filename=filePath)
        sheet = wb.active
        rowCount = sheet.max_row
        colCount = sheet.max_column
        for row in range(1, rowCount + 1):
            if row == 1:
                index = 1
                headerTexts = []
                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.componentField == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.selectedVariables = form.choices[0]
                        self.componentField = form.choices[0]
                        self.dataEntryList[formNumber]['index'] = form.choices[0]
                    else:
                        return
                else:
                    self.selectedVariables = self.componentField
                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "" and header != None:
                        item = QTreeWidgetItem(variablesItem)
                        item.setText(0, header)
                        if self.componentField == header:
                            item.setCheckState(0, Qt.Unchecked)
                        else:
                            item.setCheckState(0, Qt.Checked)
                        item.setCheckState(0, Qt.Unchecked)

                        val = str(sheet.cell(row=row + 1, column=col).value)
                        try:
                            tmp = float(val)
                            item.setCheckState(0, Qt.Checked)
                        except:
                            item.setDisabled(True)

                    if header == self.componentField:
                        componentIndex = index
                    index = index + 1
            else:
                if not sheet.cell(row=row, column=componentIndex).value in components:
                    components.append(str(sheet.cell(row=row, column=componentIndex).value))
        for component in components:
            item = QTreeWidgetItem(componentsItem)
            item.setText(0, component)
            item.setCheckState(0, Qt.Checked)
        self.components1 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file completed")

    def readXLSile(self, filePath, formNumber):
        formData = self.dataEntryList[formNumber]
        components = []
        componentIndex = 0
        selectAllItem = formData['twDynamic'].topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = formData['twDynamic'].topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file...")

        wb = xlrd.open_workbook(filePath)
        sheet = wb.sheet_by_index(0)
        rowCount = sheet.nrows
        colCount = sheet.ncols
        for row in range(rowCount):
            if row == 0:
                index = 0
                headerTexts = []
                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.componentField == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.selectedVariables = form.choices[0]
                        self.componentField = form.choices[0]
                        self.dataEntryList[formNumber]['index'] = form.choices[0]
                    else:
                        return
                else:
                    self.selectedVariables = self.componentField

                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "" and header != None:
                        item = QTreeWidgetItem(variablesItem)
                        item.setText(0, header)
                        if self.componentField == header:
                            item.setCheckState(0, Qt.Unchecked)
                        else:
                            item.setCheckState(0, Qt.Checked)
                        item.setCheckState(0, Qt.Unchecked)

                        val = str(sheet.cell_value(row, col))
                        try:
                            tmp = float(val)
                            item.setCheckState(0, Qt.Checked)
                        except:
                            item.setDisabled(True)

                    if header == self.componentField:
                        componentIndex = index
                    index = index + 1
            else:
                if not sheet.cell_value(row, componentIndex) in components:
                    components.append(sheet.cell_value(row, componentIndex))
        for component in components:
            item = QTreeWidgetItem(componentsItem)
            item.setText(0, component)
            item.setCheckState(0, Qt.Checked)
        self.components1 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file completed")

    def readCSVFile(self, filePath, formNumber):
        formData = self.dataEntryList[formNumber]
        components = []
        componentIndex = ""
        selectAllItem = formData['twDynamic'].topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = formData['twDynamic'].topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file...")

        csv_header_1 = pd.read_csv(filePath, sep=";")
        csv_header_2 = pd.read_csv(filePath)
        csv_reader = pd.read_csv(filePath, sep=";") if len(csv_header_1.columns) >= len(csv_header_2.columns) else pd.read_csv(filePath)
        headers = csv_reader.columns

        headerTexts = []

        for header in headers:
            if header != "":
                headerTexts.append(header)
        form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
        if self.componentField == "":
            if form.exec_() == QDialog.Accepted:
                headerTexts = form.choices
                self.selectedVariables = form.choices[0]
                self.componentField = form.choices[0]
                self.dataEntryList[formNumber]['index'] = form.choices[0]
                print(self.dataEntryList[formNumber]['index'])
            else:
                return
        else:
            self.selectedVariables = self.componentField

        for header in headers:
            if header != "" and header != None:
                item = QTreeWidgetItem(variablesItem)
                item.setText(0, header)
                if self.componentField == header:
                    item.setCheckState(0, Qt.Unchecked)
                else:
                    item.setCheckState(0, Qt.Checked)
                item.setCheckState(0, Qt.Unchecked)

                val = csv_reader[header].dtypes
                if val == object:
                    item.setDisabled(True)
                else:
                    item.setCheckState(0, Qt.Checked)

            if header == self.componentField:
                componentIndex = header

        cnt = len(csv_reader)
        print(cnt)
        for i in range(0, cnt):
            if not csv_reader[componentIndex][i] in components:
                print(csv_reader[componentIndex][i])
                components.append(csv_reader[componentIndex][i])

        for component in components:
            item = QTreeWidgetItem(componentsItem)
            item.setText(0, str(component))
            item.setCheckState(0, Qt.Checked)

            val = csv_reader[componentIndex].dtypes
            if val == object:
                item.setDisabled(True)
            else:
                item.setCheckState(0, Qt.Checked)
        self.components1 = components

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file completed")

    @pyqtSlot()
    def btnGenerate_clicked(self):
        print("Generating Result")
        
        ## Check if results folder is empty
        
        results_path = Path('../HTML_ASSETS/results/').resolve()
        shutil.rmtree(Path('../HTML_ASSETS/results/old/').resolve())
        os.mkdir(Path('../HTML_ASSETS/results/old/').resolve())
        for path in results_path.rglob('*.html'):
            print(path.name)
            path.rename('../HTML_ASSETS/results/old/'+path.name)

        for path in results_path.rglob('*.pdf'):
            print(path.name)
            path.rename('../HTML_ASSETS/results/old/'+path.name)
        
        if os.path.isdir(Path('../HTML_ASSETS/results/store_vars').resolve()):
            Path('../HTML_ASSETS/results/store_vars').rename('../HTML_ASSETS/results/old/store_vars')
        # try:
        #     dirContents = os.listdir(results_path)
        # except:
        #     os.mkdir(results_path)
        #     dirContents = os.listdir(results_path)

        # if not len(dirContents) == 0:
        #     msgBox = QMessageBox()
        #     deleteButton = msgBox.addButton(self.tr("Delete"), QMessageBox.ActionRole)
        #     abortButton = msgBox.addButton(QMessageBox.Abort)

        #     msgBox.setText("The result folder is not empty.")
        #     msgBox.setInformativeText("Do you want to continue deleting the previous results?")

        #     msgBox.exec_()
        #     if msgBox.clickedButton() == deleteButton:
        #             shutil.rmtree(results_path)

        #     else:
        #         return
        #     os.mkdir(results_path)
        

        self.pteLog.setPlainText(self.pteLog.toPlainText() + f"\nGenerating Result...")
        output = self.performStep1()

        if self.coordinateFilePath.text()!="":
            coordinatesFile = pd.read_csv(self.coordinateFilePath.text(), index_col=self.coordinatesIdentifier)

            coordinatesFile = coordinatesFile[[self.latColumn, self.longColumn]]


            final_df = pd.merge(output, coordinatesFile, left_index=True, right_index=True)
            final_df.rename({self.latColumn: 'lat', self.longColumn: 'long'}, axis=1, inplace=True)
            final_df = final_df.round(decimals=3)
            create_map(final_df, region_coordinates[self.regionCB.currentText().lower()])
        else:
            final_df = output
            final_df['lat'] = ""
            final_df['long'] = ""

        
        final_df.to_csv('final.csv', index=True)
        self.pteLog.setPlainText(self.pteLog.toPlainText() + f"\nFile saved to: {'final.csv'}")

        

        # output = pd.read_csv('final.csv', index_col=0)
        input_data(final_df[final_df.columns[:-2]]) # removing lat long columns

        self.pteLog.setPlainText(self.pteLog.toPlainText() + f"\nResults generated.")



    @pyqtSlot()
    def btnShow_clicked(self):
        print("Show Result")
        
        os.startfile(Path('../HTML_ASSETS/results/assets_table.html').resolve())
        
        os.startfile(Path('../HTML_ASSETS/results/main_clustering.html').resolve())

        assets_location_file = Path('../HTML_ASSETS/results/assets_location.html').resolve()
        if os.path.isfile(assets_location_file):
            os.startfile(assets_location_file)
    def performStep1(self):
        columnsList = []
        index_column = []
        for entryDict in self.dataEntryList:
            columns = []
            for i in range(entryDict['twDynamic'].topLevelItem(0).childCount()):
                column = entryDict['twDynamic'].topLevelItem(0).child(i)
                if column.checkState(0) == Qt.Checked:
                    columns.append(column.text(0))

            new_column_name = entryDict['newColumnEdit'].currentText()
            weights = entryDict['dimensionLineEdit'].text().split(',')
            weights = list(map(float, weights))
            dataset = read_file1(entryDict['txtFilePath'].text(), col_name = entryDict['index'])
            
            print(dataset)
            index_column = dataset.index.values.tolist()
            print(index_column)
            data = sum(weights, dataset[columns])
            columnsList.append(pd.DataFrame(data, columns=[new_column_name]))

        final_df = pd.concat(columnsList, axis=1)
        final_df.insert(0, 'index', index_column)
        final_df.set_index('index', inplace=True)
        final_df['Total_Indicator'] = final_df.sum(axis=1)/len(final_df.columns)


        
        return final_df.round(decimals=3)
        

            
def sum(weight, dataset):
    weights = pd.Series(weight, index=dataset.columns)
    new_data = dataset * weights
    new_data = new_data.sum(axis=1)
    return new_data

def read_file1(path, col_name):
    csv_header_1 = pd.read_csv(path, sep=";")
    csv_header_2 = pd.read_csv(path)
    df = pd.read_csv(path, index_col=col_name, sep=";") if len(csv_header_1.columns) >= len(csv_header_2.columns) else pd.read_csv(path, index_col=col_name)

    df= df.select_dtypes(include=np.number)
    NormalizeData= lambda x: (x- x.min(axis=0)) / (x.max(axis=0) - x.min(axis=0))
    data = df.apply(NormalizeData)
    return data

def main():
    app = QApplication(sys.argv)
    ex = Example()
    app.setStyleSheet(qdarkstyle.load_stylesheet())
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()