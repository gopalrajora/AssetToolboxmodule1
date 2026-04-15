import threading
import subprocess
import qdarkstyle
import shutil

from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2.QtWidgets import *

import PySide2.QtGui as QtGui
import PySide2.QtCore as QtCore
import PySide2.QtWidgets as QtWidgets

import xlrd
from openpyxl import load_workbook

from PyQt5.QtCore import pyqtSignal, pyqtSlot
import json

import chart.views
from chart.lib.main_cluster_maker import *
import numpy as np


class ChecklistDialog(QtWidgets.QDialog):

    def __init__(
            self,
            name,
            stringlist=None,
            checked=False,
            icon=None,
            parent=None,
    ):
        super(ChecklistDialog, self).__init__(parent)

        self.name = name
        self.icon = icon
        self.model = QtGui.QStandardItemModel()
        self.listView = QListView()

        for string in stringlist:
            item = QtGui.QStandardItem(string)
            # item.setCheckable(True)
            check = \
                (QtCore.Qt.Checked if checked else QtCore.Qt.Unchecked)
            # item.setCheckState(check)
            self.model.appendRow(item)

        self.listView.setModel(self.model)
        self.listView.setCurrentIndex(self.model.index(0, 0));

        self.okButton = QPushButton('OK')
        self.cancelButton = QPushButton('Cancel')

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(self.okButton)
        hbox.addWidget(self.cancelButton)

        vbox = QVBoxLayout(self)
        vbox.addWidget(self.listView)
        vbox.addLayout(hbox)

        self.setWindowTitle(self.name)
        if self.icon:
            self.setWindowIcon(self.icon)

        self.okButton.clicked.connect(self.onAccepted)
        self.cancelButton.clicked.connect(self.reject)

    def onAccepted(self):
        if self.listView.selectedIndexes().__len__() != 1:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText('Please select only one field')
            msg.setWindowTitle("Error")
            msg.exec_()
            return
        self.choices = [self.listView.selectedIndexes()[0].data()]
        if self.choices.__len__() != 1:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText('Please select only one field')
            msg.setWindowTitle("Error")
            msg.exec_()
        else:
            self.accept()


class Example(QtWidgets.QDialog):

    def __init__(self):
        super().__init__()

        self.initUI()

        self.selectedVariables = ""

    def initUI(self):
        self.totalVLayout = QVBoxLayout(self)
        self.totalVLayout.setObjectName(u"totalVLayout")
        self.totalVLayout.setContentsMargins(20, 20, 20, 20)
        self.totalVLayout.setSpacing(15)

        self.logoTitleHLayout = QHBoxLayout()
        self.logoTitleHLayout.setSpacing(10)
        self.logoTitleHLayout.setObjectName(u"fileBrowse1HLayout")

        self.lblLogo = QLabel(self)
        pixmap = QPixmap('static/resources/logo_white.png')
        self.lblLogo.setPixmap(pixmap)
        self.lblLogo.setScaledContents(True)
        self.lblLogo.setMaximumHeight(50)
        self.lblLogo.setMaximumWidth(300)
        self.logoTitleHLayout.addWidget(self.lblLogo)

        self.lblTitle = QLabel(self)
        self.lblTitle.setText("Condition Characterization")
        self.lblTitle.setAlignment(Qt.AlignCenter)
        self.lblTitle.setFont(QFont("Arial", 18))
        self.logoTitleHLayout.addWidget(self.lblTitle)

        self.totalVLayout.addLayout(self.logoTitleHLayout)

        self.assetTypeHLayout = QHBoxLayout()
        self.assetTypeHLayout.setSpacing(10)
        self.assetTypeHLayout.setObjectName(u"fileBrowse1HLayout")

        self.txtAssestType = QLineEdit(self)
        self.txtAssestType.setObjectName(u"txtFilePath")
        self.txtAssestType.setPlaceholderText("Type of Assest")
        self.txtAssestType.setMaximumSize(QSize(250, 30))

        self.assetTypeHLayout.addWidget(self.txtAssestType)

        self.lblSpace = QLabel(self)
        self.assetTypeHLayout.addWidget(self.lblSpace)
        self.totalVLayout.addLayout(self.assetTypeHLayout)

        self.firstVLayout = QVBoxLayout()
        self.firstVLayout.setSpacing(10)
        self.firstLabel = QLabel(self)
        self.firstLabel.setText(" Select the CSV input File")
        self.firstVLayout.addWidget(self.firstLabel)

        self.fileBrowse1HLayout = QHBoxLayout()
        self.fileBrowse1HLayout.setSpacing(10)
        self.fileBrowse1HLayout.setObjectName(u"fileBrowse1HLayout")

        self.txtFilePath = QLineEdit(self)
        self.txtFilePath.setObjectName(u"txtFilePath")
        self.txtFilePath.setReadOnly(True)

        self.fileBrowse1HLayout.addWidget(self.txtFilePath)

        self.btnBrowse = QPushButton(self)
        self.btnBrowse.setObjectName(u"btnBrowse")
        self.btnBrowse.setMinimumSize(QSize(80, 0))
        self.btnBrowse.setText("Browse")
        self.btnBrowse.clicked.connect(self.btnBrowse_clicked)

        self.fileBrowse1HLayout.addWidget(self.btnBrowse)

        self.firstVLayout.addLayout(self.fileBrowse1HLayout)
        self.totalVLayout.addLayout(self.firstVLayout)

        self.selectComponentVariableHLayout = QHBoxLayout()
        self.selectComponentVariableHLayout.setSpacing(15)
        self.selectComponentVariableHLayout.setObjectName(u"selectComponentVariableHLayout")

        self.secondVLayout = QVBoxLayout()
        self.secondVLayout.setSpacing(10)
        self.secondLabel = QLabel(self)
        self.secondLabel.setText("Select Components or Variables")
        self.secondVLayout.addWidget(self.secondLabel)

        self.twDynamic = QTreeWidget(self)
        __qtreewidgetitem = QTreeWidgetItem(self.twDynamic)
        __qtreewidgetitem.setCheckState(0, Qt.Checked)
        # __qtreewidgetitem1 = QTreeWidgetItem(__qtreewidgetitem)
        # __qtreewidgetitem1.setCheckState(0, Qt.Checked)
        # __qtreewidgetitem4 = QTreeWidgetItem(__qtreewidgetitem)
        # __qtreewidgetitem4.setCheckState(0, Qt.Checked)
        self.twDynamic.setObjectName(u"twDynamic")
        self.twDynamic.header().setVisible(False)
        self.twDynamic.itemClicked.connect(self.twDynamic_onItemClicked)

        self.secondVLayout.addWidget(self.twDynamic)

        self.auxLabel = QLabel(self)
        self.auxLabel.setText("Select Auxiliary file")
        self.secondVLayout.addWidget(self.auxLabel)

        self.fileBrowse2HLayout = QHBoxLayout()
        self.fileBrowse2HLayout.setSpacing(10)
        self.fileBrowse2HLayout.setObjectName(u"fileBrowse2HLayout")

        self.txtFilePath2 = QLineEdit(self)
        self.txtFilePath2.setObjectName(u"txtFilePath2")
        self.txtFilePath2.setReadOnly(True)

        self.fileBrowse2HLayout.addWidget(self.txtFilePath2)

        self.btnBrowse2 = QPushButton(self)
        self.btnBrowse2.setObjectName(u"btnBrowse")
        self.btnBrowse2.setMinimumSize(QSize(80, 0))
        self.btnBrowse2.setText("Browse")
        self.btnBrowse2.clicked.connect(self.btnBrowse2_clicked)

        self.fileBrowse2HLayout.addWidget(self.btnBrowse2)

        self.secondVLayout.addLayout(self.fileBrowse2HLayout)

        self.listVariableLabel = QLabel(self)
        self.listVariableLabel.setText("Group by Attribute")
        self.secondVLayout.addWidget(self.listVariableLabel)

        self.cmbVariables2 = QComboBox(self)
        self.cmbVariables2.setObjectName(u"cmbVariables2")
        self.secondVLayout.addWidget(self.cmbVariables2)

        self.selectComponentVariableHLayout.addLayout(self.secondVLayout)

        self.verticalLayout = QVBoxLayout()
        self.verticalLayout.setObjectName(u"verticalLayout")
        self.thirdLabel = QLabel(self)
        self.thirdLabel.setText("Select the Dimension to be saved")
        self.verticalLayout.addWidget(self.thirdLabel)

        self.cmbConfig = QComboBox(self)
        self.cmbConfig.addItem("Life Assessment")
        self.cmbConfig.addItem("Maintenance Strategy")
        self.cmbConfig.addItem("Economic Impact")
        self.cmbConfig.setObjectName(u"cmbConfig")
        self.cmbConfig.currentIndexChanged.connect(self.cmbConfig_IndexChanged)

        self.verticalLayout.addWidget(self.cmbConfig)

        self.listConfig = QListWidget(self)
        self.listConfig.setMinimumHeight(160)
        self.listConfig.setObjectName(u"listConfig")
        self.listConfig.itemDoubleClicked.connect(self.listConfig_dblClick)

        self.verticalLayout.addWidget(self.listConfig)

        self.forthLabel = QLabel(self)
        self.forthLabel.setText(" Select the Algorithm")
        self.verticalLayout.addWidget(self.forthLabel)

        self.horizontalLayoutM = QHBoxLayout()
        self.horizontalLayoutM.setSpacing(10)
        self.horizontalLayoutM.setObjectName(u"fileBrowse1HLayout")

        self.cmbMethod = QComboBox(self)
        self.cmbMethod.addItem("")
        self.cmbMethod.addItem("")
        self.cmbMethod.addItem("")
        self.cmbMethod.setObjectName(u"cmbMethod")

        self.btnClearSettings = QPushButton(self)
        self.btnClearSettings.setObjectName(u"btnClearSettings")
        self.btnClearSettings.setText("Clear")
        # self.btnClearSettings.setSizePolicy(QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed))
        self.btnClearSettings.clicked.connect(self.clearSettingsFunc)

        self.btnSaveSetting = QPushButton(self)
        self.btnSaveSetting.setObjectName(u"btnSaveSetting")
        self.btnSaveSetting.setText("Save Setting")
        self.btnSaveSetting.clicked.connect(self.btnSaveSetting_clicked)

        self.horizontalLayoutM.addWidget(self.cmbMethod)

        self.verticalLayout.addLayout(self.horizontalLayoutM)

        self.twStatic = QTreeWidget(self)
        __qtreewidgetitem8 = QTreeWidgetItem(self.twStatic)
        __qtreewidgetitem8.setCheckState(0, Qt.Checked)
        __qtreewidgetitem9 = QTreeWidgetItem(__qtreewidgetitem8)
        __qtreewidgetitem9.setCheckState(0, Qt.Checked)
        __qtreewidgetitem10 = QTreeWidgetItem(__qtreewidgetitem8)
        __qtreewidgetitem10.setCheckState(0, Qt.Checked)
        __qtreewidgetitem11 = QTreeWidgetItem(__qtreewidgetitem8)
        __qtreewidgetitem11.setCheckState(0, Qt.Checked)
        self.twStatic.setObjectName(u"twStatic")
        self.twStatic.header().setVisible(False)
        self.twStatic.itemClicked.connect(self.twStatic_onItemClicked)

        ___qtreewidgetitem = self.twDynamic.headerItem()
        ___qtreewidgetitem.setText(0, QCoreApplication.translate("Form", u"Items", None))

        self.verticalLayout.addWidget(self.twStatic)

        self.selectComponentVariableHLayout.addLayout(self.verticalLayout)

        self.totalVLayout.addLayout(self.selectComponentVariableHLayout)

        self.pteLog = QPlainTextEdit(self)
        self.pteLog.setObjectName(u"pteLog")
        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pteLog.sizePolicy().hasHeightForWidth())
        self.pteLog.setSizePolicy(sizePolicy)
        self.pteLog.setMinimumSize(QSize(0, 200))
        self.pteLog.setMaximumSize(QSize(16777215, 400))

        self.totalVLayout.addWidget(self.pteLog)

        self.horizontalLayout_31 = QHBoxLayout()
        self.horizontalLayout_31.setSpacing(20)
        self.horizontalLayout_31.setObjectName(u"horizontalLayout_31")
        self.horizontalLayout_31.setContentsMargins(20, -1, 20, -1)
        
        self.horizontalLayout_31.addWidget(self.btnClearSettings)
        self.horizontalLayout_31.addWidget(self.btnSaveSetting)

        self.horizontalLayout_3 = QHBoxLayout()
        self.horizontalLayout_3.setSpacing(20)
        self.horizontalLayout_3.setObjectName(u"horizontalLayout_3")
        self.horizontalLayout_3.setContentsMargins(20, -1, 20, -1)

        self.btnGenerate = QPushButton(self)
        self.btnGenerate.setObjectName(u"btnGenerate")
        self.btnGenerate.clicked.connect(self.btnGenerate_clicked)

        self.horizontalLayout_3.addWidget(self.btnGenerate)

        self.btnShow = QPushButton(self)
        self.btnShow.setObjectName(u"btnShow")
        self.btnShow.clicked.connect(self.btnShow_clicked)
        self.horizontalLayout_3.addWidget(self.btnShow)

        self.totalVLayout.addLayout(self.horizontalLayout_31)
        self.totalVLayout.addLayout(self.horizontalLayout_3)

        self.retranslateUi()

        self.configs = []
        self.component2Field = ""
        self.component1Field = ""
        self.components2 = []
        self.variables2 = []

        self.setGeometry(100, 100, 800, 640)
        self.setWindowTitle('WP 5.1 Tools')
        self.show()

        # self.LoadSetting()

    def retranslateUi(self):
        __sortingEnabled = self.twDynamic.isSortingEnabled()
        self.twDynamic.setSortingEnabled(False)
        ___qtreewidgetitem1 = self.twDynamic.topLevelItem(0)
        ___qtreewidgetitem1.setText(0, QCoreApplication.translate("Form", u"Variables", None))
        # ___qtreewidgetitem2 = ___qtreewidgetitem1.child(0)
        # ___qtreewidgetitem2.setText(0, QCoreApplication.translate("Form", u"Components", None))
        # ___qtreewidgetitem5 = ___qtreewidgetitem1.child(1)
        # ___qtreewidgetitem5.setText(0, QCoreApplication.translate("Form", u"Variables", None))
        self.twDynamic.setSortingEnabled(__sortingEnabled)

        self.cmbMethod.setItemText(0, QCoreApplication.translate("Form", u"Auto", None))
        self.cmbMethod.setItemText(1, QCoreApplication.translate("Form", u"Elbow Method", None))
        self.cmbMethod.setItemText(2, QCoreApplication.translate("Form", u"Gap Statistics", None))

        ___qtreewidgetitem8 = self.twStatic.headerItem()
        ___qtreewidgetitem8.setText(0, QCoreApplication.translate("Form", u"Items", None))

        __sortingEnabled1 = self.twStatic.isSortingEnabled()
        self.twStatic.setSortingEnabled(False)
        ___qtreewidgetitem10 = self.twStatic.topLevelItem(0)
        ___qtreewidgetitem10.setText(0, QCoreApplication.translate("Form", u"Result", None))
        ___qtreewidgetitem11 = ___qtreewidgetitem10.child(0)
        ___qtreewidgetitem11.setText(0, QCoreApplication.translate("Form", u"Asset assessment", None));
        ___qtreewidgetitem12 = ___qtreewidgetitem10.child(1)
        ___qtreewidgetitem12.setText(0, QCoreApplication.translate("Form", u"Number of Clusters assessment", None));
        ___qtreewidgetitem13 = ___qtreewidgetitem10.child(2)
        ___qtreewidgetitem13.setText(0, QCoreApplication.translate("Form", u"Clustering Training process", None));
        self.twStatic.setSortingEnabled(__sortingEnabled1)

        self.btnGenerate.setText(QCoreApplication.translate("Form", u"Generate Result", None))
        self.btnShow.setText(QCoreApplication.translate("Form", u"Show Result", None))


    @pyqtSlot(QtWidgets.QTreeWidgetItem, int)
    def twDynamic_onItemClicked(self, it, col):
        item_count = it.childCount()
        for i in range(item_count):
            child_item = it.child(i)
            if child_item.isDisabled():
                continue
            child_item.setCheckState(0, it.checkState(0))
            for j in range(child_item.childCount()):
                item = child_item.child(j)
                if item.isDisabled():
                    continue
                item.setCheckState(0, it.checkState(0))

        variablesItem = self.twDynamic.topLevelItem(0)
        variablesItem.setCheckState(0, Qt.Checked)

        item_count = variablesItem.childCount()
        for i in range(item_count):
            child_item = variablesItem.child(i)
            if child_item.isDisabled():
                continue
            if child_item.parent() != None:
                if variablesItem.checkState(0) == Qt.Checked:
                    variablesItem.setCheckState(0, child_item.checkState(0))


    @pyqtSlot(QtWidgets.QTreeWidgetItem, int)
    def twStatic_onItemClicked(self, it, col):
        item_count = it.childCount()
        for i in range(item_count):
            child_item = it.child(i)
            child_item.setCheckState(0, it.checkState(0))
            for j in range(child_item.childCount()):
                item = child_item.child(j)
                item.setCheckState(0, it.checkState(0))

        ResultItem = self.twStatic.topLevelItem(0)
        ResultItem.setCheckState(0, Qt.Checked)

        item_count = ResultItem.childCount()
        for i in range(item_count):
            child_item = ResultItem.child(i)
            if child_item.parent() != None:
                if ResultItem.checkState(0) == Qt.Checked:
                    ResultItem.setCheckState(0, child_item.checkState(0))

    @pyqtSlot()
    def clearSettingsFunc(self):
        if os.path.isfile("./config.cfg"):
            os.remove("./config.cfg")
        self.listConfig.clear()
        self.configs = []
        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Clear Settings")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nClear Setting")

    def closeEvent(self, event):
        if os.path.isfile("./config.cfg"):
            os.remove("./config.cfg")

    @pyqtSlot()
    def btnBrowse_clicked(self):
        qfd = QFileDialog()
        path = ""
        filter = "Excel Files(*.xls *.xlsx *.xlsm *.csv)"
        title = "Open file"
        f = QFileDialog.getOpenFileName(qfd, title, path, filter)
        print(f)
        if f[0] != "":
            self.txtFilePath.setText(f[0])
        else:
            return
        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Browse file")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nBrowse file")

        filePath = self.txtFilePath.text()
        self.clearDynamicTree()

        filename, file_extension = os.path.splitext(filePath)
        print(file_extension)
        if file_extension == ".csv":
            self.readCSVFile(filePath)
        elif file_extension == ".xls":
            self.readXLSile(filePath)
        else:
            self.readXLSXile(filePath)

    @pyqtSlot()
    def btnBrowse2_clicked(self):
        qfd = QFileDialog()
        path = ""
        filter = "Excel Files(*.xls *.xlsx *.xlsm *.csv)"
        title = "Open file"
        f = QFileDialog.getOpenFileName(qfd, title, path, filter)
        print(f)
        if f[0] != "":
            self.txtFilePath2.setText(f[0])
        else:
            return

        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Browse file")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nBrowse file")

        filePath = self.txtFilePath2.text()
        self.clearPart2()

        filename, file_extension = os.path.splitext(filePath)
        print(file_extension)
        if file_extension == ".csv":
            self.readCSVFile2(filePath)
        elif file_extension == ".xls":
            self.readXLSile2(filePath)
        else:
            self.readXLSXile2(filePath)

    @pyqtSlot()
    def btnSaveSetting_clicked(self):
        # selectAllItem = self.twDynamic.topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = self.twDynamic.topLevelItem(0)

        variables = []
        selectedVariables = self.selectedVariables
        components = []
        results = []

        # for i in range(componentsItem.childCount()):
        #     if componentsItem.child(i).checkState(0) == Qt.Checked:
        #         print(componentsItem.child(i).text(0))
        #         components.append(componentsItem.child(i).text(0))
        for i in range(variablesItem.childCount()):
            # selectedVariables.append(variablesItem.child(i).text(0))
            if variablesItem.child(i).checkState(0) == Qt.Checked:
                print(variablesItem.child(i).text(0))
                variables.append(variablesItem.child(i).text(0))

        method = self.cmbMethod.currentText()
        resultItem = self.twStatic.topLevelItem(0)
        for i in range(resultItem.childCount()):
            if resultItem.child(i).checkState(0) == Qt.Checked:
                results.append(resultItem.child(i).text(0))

        flag = True
        for i in range(self.configs.__len__()):
            if self.configs[i]["config"] == self.cmbConfig.currentText():
                self.configs[i]["path"] = self.txtFilePath.text()
                self.configs[i]["variables"] = variables
                self.configs[i]["components"] = self.components1
                self.configs[i]["selectedVariables"] = selectedVariables
                self.configs[i]["method"] = method
                self.configs[i]["results"] = results
                self.configs[i]["path2"] = self.txtFilePath2.text()
                self.configs[i]["variables2"] = self.cmbVariables2.currentText()
                self.configs[i]["components2"] = self.components2
                self.configs[i]["component2_field"] = self.component2Field
                self.configs[i]["component1_field"] = self.component1Field
                self.configs[i]["assestsType"] = self.txtAssestType.text()
                flag = False
        if flag:
            self.configs.append({
                "config": self.cmbConfig.currentText(),
                "path": self.txtFilePath.text(),
                "selectedVariables": selectedVariables,
                "variables": variables,
                "component1_field": self.component1Field,
                "components": self.components1,
                "method": method,
                "results": results,
                "path2": self.txtFilePath2.text(),
                "variables2": self.cmbVariables2.currentText(),
                "components2": self.components2,
                "component2_field": self.component2Field,
                "assestsType": self.txtAssestType.text()
            })
            self.listConfig.addItem(self.cmbConfig.currentText())

        # qfd = QFileDialog()
        # path = ""
        # filter = "Config File(*.cfg)"
        # title = "Save Config"
        # path = QFileDialog.getSaveFileName(qfd, title, path, filter)
        # print(path)
        path = "config.cfg"
        if path != "":
            with open(path, 'w') as f:
                json.dump(self.configs, f)
        if self.pteLog.toPlainText() == "":
            self.pteLog.setPlainText("Save Settings")
        else:
            self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nSave Setting")

    
    @pyqtSlot(int)
    def cmbConfig_IndexChanged(self, index):
        self.loadSetting(self.cmbConfig.currentText())        

    @pyqtSlot(QtWidgets.QListWidgetItem)
    def listConfig_dblClick(self, item):
        self.loadSetting(item.text())

    def loadSetting(self, settingName):
        exists = False
        for i in range(self.configs.__len__()):
            if self.configs[i]["config"] == settingName:
                exists = True
                self.txtFilePath.setText(self.configs[i]["path"])
                self.txtFilePath2.setText(self.configs[i]["path2"])
                filePath = self.txtFilePath.text()
                self.clearDynamicTree()

                filename, file_extension = os.path.splitext(filePath)
                print(file_extension)
                self.component1Field = self.configs[i]["component1_field"]
                if filePath != "":
                    if file_extension == ".csv":
                        self.readCSVFile(filePath)
                    elif file_extension == ".xls":
                        self.readXLSile(filePath)
                    elif file_extension == ".xlsx":
                        self.readXLSXile(filePath)

                filePath = self.txtFilePath2.text()
                filename, file_extension = os.path.splitext(filePath)
                self.clearPart2()
                self.component2Field = self.configs[i]["component2_field"]
                print(file_extension)
                if filePath != "":
                    if file_extension == ".csv":
                        self.readCSVFile2(filePath)
                    elif file_extension == ".xls":
                        self.readXLSile2(filePath)
                    elif file_extension == ".xlsx":
                        self.readXLSXile2(filePath)
                variablesItem = self.twDynamic.topLevelItem(0)

                self.cmbMethod.setCurrentText(self.configs[i]["method"])

                for j in range(variablesItem.childCount()):
                    if variablesItem.child(j).text(0) in self.configs[i]["variables"]:
                        variablesItem.child(j).setCheckState(0, Qt.Checked)
                    else:
                        variablesItem.child(j).setCheckState(0, Qt.Unchecked)

                resultItem = self.twStatic.topLevelItem(0)
                for j in range(resultItem.childCount()):
                    if resultItem.child(j).text(0) in self.configs[i]["results"]:
                        resultItem.child(j).setCheckState(0, Qt.Checked)
                    else:
                        resultItem.child(j).setCheckState(0, Qt.Unchecked)

                self.txtAssestType.setText(self.configs[i]["assestsType"])
                self.component2Field = self.configs[i]["component2_field"]
                self.components2 = self.configs[i]["components2"]
                self.cmbVariables2.setCurrentText(self.configs[i]["variables2"])

                break

        variablesItem = self.twDynamic.topLevelItem(0)
        variablesItem.setCheckState(0, Qt.Checked)

        item_count = variablesItem.childCount()
        for i in range(item_count):
            child_item = variablesItem.child(i)
            if child_item.parent() != None:
                if variablesItem.checkState(0) == Qt.Checked:
                    variablesItem.setCheckState(0, child_item.checkState(0))
        
        if not exists:
            self.clearPart2()
            self.txtFilePath2.setText("")
            self.cmbVariables2.clear()

    def clearDynamicTree(self):
        # selectAllItem = self.twDynamic.topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = self.twDynamic.topLevelItem(0)

        # for i in reversed(range(componentsItem.childCount())):
        #     componentsItem.removeChild(componentsItem.child(i))
        for i in reversed(range(variablesItem.childCount())):
            variablesItem.removeChild(variablesItem.child(i))
        self.component1Field = ""

    def clearPart2(self):
        self.cmbVariables2.clear()
        self.component2Field = ""
        self.components2 = []
        self.variables2 = []

    def readXLSXile(self, filePath):
        components = []
        componentIndex = 0
        # selectAllItem = self.twDynamic.topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = self.twDynamic.topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file...")
        wb = load_workbook(filename=filePath)
        sheet = wb.active
        rowCount = sheet.max_row
        colCount = sheet.max_column
        for row in range(1, rowCount + 1):
            if row == 1:
                index = 1
                headerTexts = []
                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.component1Field == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.selectedVariables = form.choices[0]
                        self.component1Field = form.choices[0]
                    else:
                        return
                else:
                    self.selectedVariables = self.component1Field
                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "" and header != None:
                        item = QTreeWidgetItem(variablesItem)
                        item.setText(0, header)
                        if self.component1Field == header:
                            item.setCheckState(0, Qt.Unchecked)
                        else:
                            item.setCheckState(0, Qt.Checked)
                        item.setCheckState(0, Qt.Unchecked)

                        val = str(sheet.cell(row=row + 1, column=col).value)
                        try:
                            tmp = float(val)
                            item.setCheckState(0, Qt.Checked)
                        except:
                            item.setDisabled(True)

                    if header == self.component1Field:
                        componentIndex = index
                    index = index + 1
            else:
                if not sheet.cell(row=row, column=componentIndex).value in components:
                    components.append(str(sheet.cell(row=row, column=componentIndex).value))
        # for component in components:
        #     item = QTreeWidgetItem(componentsItem)
        #     item.setText(0, component)
        #     item.setCheckState(0, Qt.Checked)
        self.components1 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file completed")

    def readXLSile(self, filePath):
        components = []
        componentIndex = 0
        # selectAllItem = self.twDynamic.topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = self.twDynamic.topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file...")

        wb = xlrd.open_workbook(filePath)
        sheet = wb.sheet_by_index(0)
        rowCount = sheet.nrows
        colCount = sheet.ncols
        for row in range(rowCount):
            if row == 0:
                index = 0
                headerTexts = []
                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.component1Field == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.selectedVariables = form.choices[0]
                        self.component1Field = form.choices[0]
                    else:
                        return
                else:
                    self.selectedVariables = self.component1Field

                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "" and header != None:
                        item = QTreeWidgetItem(variablesItem)
                        item.setText(0, header)
                        if self.component1Field == header:
                            item.setCheckState(0, Qt.Unchecked)
                        else:
                            item.setCheckState(0, Qt.Checked)
                        item.setCheckState(0, Qt.Unchecked)

                        val = str(sheet.cell_value(row, col))
                        try:
                            tmp = float(val)
                            item.setCheckState(0, Qt.Checked)
                        except:
                            item.setDisabled(True)

                    if header == self.component1Field:
                        componentIndex = index
                    index = index + 1
            else:
                if not sheet.cell_value(row, componentIndex) in components:
                    components.append(sheet.cell_value(row, componentIndex))
        # for component in components:
        #     item = QTreeWidgetItem(componentsItem)
        #     item.setText(0, component)
        #     item.setCheckState(0, Qt.Checked)
        self.components1 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file completed")

    def readCSVFile(self, filePath):
        components = []
        componentIndex = ""
        # selectAllItem = self.twDynamic.topLevelItem(0)
        # componentsItem = selectAllItem.child(0)
        variablesItem = self.twDynamic.topLevelItem(0)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file...")

        csv_header_1 = pd.read_csv(filePath, sep=";",index_col=0, nrows=0)
        csv_header_2 = pd.read_csv(filePath, index_col=0, nrows=0)
        csv_reader = pd.read_csv(filePath, sep=";") if csv_header_1.columns.__len__() >= csv_header_2.columns.__len__() else pd.read_csv(filePath)

        headers = csv_reader.columns

        headerTexts = []

        for header in headers:
            if header != "":
                headerTexts.append(header)
        form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
        if self.component1Field == "":
            if form.exec_() == QDialog.Accepted:
                headerTexts = form.choices
                self.selectedVariables = form.choices[0]
                self.component1Field = form.choices[0]
            else:
                return
        else:
            self.selectedVariables = self.component1Field

        for header in headers:
            if header != "" and header != None:
                item = QTreeWidgetItem(variablesItem)
                item.setText(0, header)
                if self.component1Field == header:
                    item.setCheckState(0, Qt.Unchecked)
                else:
                    item.setCheckState(0, Qt.Checked)
                item.setCheckState(0, Qt.Unchecked)

                val = csv_reader[header].dtypes
                if val == object:
                    item.setDisabled(True)
                else:
                    item.setCheckState(0, Qt.Checked)

            if header == self.component1Field:
                componentIndex = header

        cnt = len(csv_reader)
        print(cnt)
        for i in range(0, cnt):
            if not csv_reader[componentIndex][i] in components:
                # print(csv_reader[componentIndex][i])
                components.append(csv_reader[componentIndex][i])

        # for component in components:
        #     item = QTreeWidgetItem(componentsItem)
        #     item.setText(0, str(component))
        #     item.setCheckState(0, Qt.Checked)

        #     val = csv_reader[componentIndex].dtypes
        #     if val == object:
        #         item.setDisabled(True)
        #     else:
        #         item.setCheckState(0, Qt.Checked)
        self.components1 = components

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file completed")

    def readXLSXile2(self, filePath):
        components = []
        componentIndex = 0
        selectAllItem = self.twDynamic.topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = selectAllItem.child(1)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file...")
        wb = load_workbook(filename=filePath)
        sheet = wb.active
        rowCount = sheet.max_row
        colCount = sheet.max_column
        for row in range(1, rowCount + 1):
            if row == 1:
                index = 1
                headerTexts = []
                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.component2Field == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.component2Field = form.choices[0]
                    else:
                        return

                for col in range(1, colCount + 1):
                    header = str(sheet.cell(row=row, column=col).value)
                    if header != "" and header != None:
                        self.variables2.append(header)
                        self.cmbVariables2.addItem(str(header))

                    if header == self.component2Field:
                        componentIndex = index
                        self.component2Field = header
                    index = index + 1
            else:
                if not sheet.cell(row=row, column=componentIndex).value in components:
                    components.append(str(sheet.cell(row=row, column=componentIndex).value))

        self.components2 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xlsx file completed")

    def readXLSile2(self, filePath):
        components = []
        componentIndex = 0
        selectAllItem = self.twDynamic.topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = selectAllItem.child(1)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file...")

        wb = xlrd.open_workbook(filePath)
        sheet = wb.sheet_by_index(0)
        rowCount = sheet.nrows
        colCount = sheet.ncols
        for row in range(rowCount):
            if row == 0:
                index = 0
                headerTexts = []
                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "":
                        headerTexts.append(header)
                form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
                if self.component2Field == "":
                    if form.exec_() == QDialog.Accepted:
                        headerTexts = form.choices
                        self.component2Field = form.choices[0]
                    else:
                        return

                for col in range(colCount):
                    header = str(sheet.cell_value(row, col))
                    if header != "" and header != None:
                        self.variables2.append(header)
                        self.cmbVariables2.addItem(str(header))

                    if header == self.component2Field:
                        componentIndex = index
                        self.component2Field = header
                    index = index + 1
            else:
                if not sheet.cell_value(row, componentIndex) in components:
                    components.append(sheet.cell_value(row, componentIndex))
        self.components2 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading xls file completed")

    def readCSVFile2(self, filePath):
        components = []
        componentIndex = ""
        selectAllItem = self.twDynamic.topLevelItem(0)
        componentsItem = selectAllItem.child(0)
        variablesItem = selectAllItem.child(1)

        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file...")

        csv_header_1 = pd.read_csv(filePath, sep=";", decimal=",", index_col=0, nrows=0)
        csv_header_2 = pd.read_csv(filePath, index_col=0, nrows=0)
        csv_reader = pd.read_csv(filePath, sep=";", decimal=",") if csv_header_1.columns.__len__() >= csv_header_2.columns.__len__() else pd.read_csv(filePath)

        headers = csv_reader.columns

        headerTexts = []

        for header in headers:
            if header != "":
                headerTexts.append(header)
        form = ChecklistDialog('Select Header Columns', headerTexts, checked=False)
        if self.component2Field == "":
            if form.exec_() == QDialog.Accepted:
                headerTexts = form.choices
                self.component2Field = form.choices[0]
            else:
                return

        for header in headers:
            if header != "" and header != None:
                self.variables2.append(header)
                self.cmbVariables2.addItem(str(header))

            if header == self.component2Field:
                componentIndex = header
                self.component2Field = header

        cnt = len(csv_reader)
        for i in range(0, cnt):
            if not csv_reader[componentIndex][i] in components:
                components.append(csv_reader[componentIndex][i])

        self.components2 = components
        self.pteLog.setPlainText(self.pteLog.toPlainText() + "\nReading csv file completed")

    @pyqtSlot()
    def btnGenerate_clicked(self):
        print("Generating Result")
        ## Check if results folder is empty
        results_path = './static/results_condition_characterization/'
        clustering_process_folder = "clustering_graphs"
        restore_folder = "restore_files"
        try:
            new_folder_created = False
            dirContents = os.listdir(results_path)
        except:
            os.mkdir(results_path)
            os.mkdir(results_path + f"{clustering_process_folder}/")
            os.mkdir(results_path + f"{restore_folder}/")
            new_folder_created = True
            dirContents = os.listdir(results_path)

        if not len(dirContents) == 0 and not new_folder_created:
            if not clustering_process_folder in dirContents:
                os.mkdir(results_path + f"{clustering_process_folder}/")
            if not restore_folder in dirContents:
                os.mkdir(results_path + f"{restore_folder}/")

            msgBox = QMessageBox()
            deleteButton = msgBox.addButton(self.tr("Delete"), QMessageBox.ActionRole)
            abortButton = msgBox.addButton(QMessageBox.Abort)

            msgBox.setText("The result folder is not empty.")
            msgBox.setInformativeText("Do you want to continue deleting the previous results?")

            msgBox.exec_()
            if msgBox.clickedButton() == deleteButton:
                try:
                    shutil.rmtree(results_path)
                    os.mkdir(results_path)
                    os.mkdir(results_path+ f"{clustering_process_folder}/")
                    os.mkdir(results_path+ f"{restore_folder}/")
                except:
                    pass

                input_data(self.configs)
                print("Done")

            elif msgBox.clickedButton() == abortButton:
                print("Abort")

        else:
            if not clustering_process_folder in dirContents:
                os.mkdir(results_path + f"{clustering_process_folder}/")

            if not restore_folder in dirContents:
                os.mkdir(results_path + f"{restore_folder}/")
            input_data(self.configs)
            print("Done")





    @pyqtSlot()
    def btnShow_clicked(self):
        open_main = True
        result_path = 'static/results_condition_characterization/'

        types_standard = {'Economic Impact':"economic_impact",
                          'Maintenance Strategy':"maintenance_strategy",
                          'Life Assessment':"life_assessment",
                          'Number of Clusters assessment': "number_clusters",
                          'Clustering Training process': "clustering_process"} ## NOT MODIF

        print("Show Result")
        subprocess.Popen("python Scripts/manage.py runserver --noreload")


#       for i in range(self.configs.__len__()):
#           for result_type in self.configs[i]["results"]:
#               if result_type != 'Asset assessment' and open_main:
#                   path_file = result_path + types_standard[self.configs[i]["config"]]+"_"+types_standard[result_type]+".pdf"
#                   os.startfile(path_file.replace("/", "\\"))



def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet())
    ex = Example()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()


